import openpyxl
import os
from django.core.management.base import BaseCommand
from jewelry_app.models import Category, Product, ProductImage, Stone
from django.contrib.auth.models import User

# Mapping артикулов к категориям
ARTICLE_CATEGORY_MAP = {
    'RG': 'Кольца',
    'ER': 'Серьги',
    'BR': 'Браслеты',
    'PN': 'Кулоны',
    'CH': 'Цепочки',
    'NC': 'Ожерелья',
    'ST': 'Наборы',
    'FBR': 'Ножные браслеты',
}

# Mapping камней из Excel
STONE_MAP = {
    'PPL': 'Жемчуг',
    'AMT': 'Аметист',
    'TRQ': 'Бирюза',
    'TPZ': 'Топаз',
    'CRL': 'Коралл',
    'EMD': 'Изумруд',
    'QRT': 'Кварц',
    'RBY': 'Рубин',
    'AMB': 'Янтарь',
    'DMD': 'Бриллиант',
    'PRD': 'Перидот',
    'GRN': 'Гранат',
    'SPP': 'Сапфир',
    'JSP': 'Яшма',
    'OBS': 'Обсидиан',
    'LAP': 'Лазурит',
    'MAL': 'Малахит',
    'AGT': 'Агат',
    'ONY': 'Оникс',
    'CTN': 'Цитрин',
    'AQU': 'Аквамарин',
    'OPAL': 'Опал',
}

# Красивые названия
PRODUCT_NAMES = {
    'Кольца': ['Вечность', 'Грация', 'Империя', 'Нежность', 'Страсть', 'Аурелия', 'Белла', 'Виктория', 'Гармония', 'Мечта', 'Сияние', 'Афродита', 'Луна', 'Солнце', 'Звезда', 'Комета'],
    'Серьги': ['Луна', 'Звезда', 'Каприз', 'Мелодия', 'Поэзия', 'Аделия', 'Бриллиант', 'Венера', 'Глория', 'Джулия', 'Элизабет', 'Флора', 'Хлоя', 'Роза', 'Ночь', 'Утро'],
    'Браслеты': ['Шарм', 'Грация', 'Фортуна', 'Мечта', 'Гармония', 'Афина', 'Бриллиант', 'Волна', 'Ветер', 'Океан', 'Река', 'Горы', 'Лес', 'Пустыня'],
    'Кулоны': ['Сердце', 'Солнце', 'Звезда', 'Луна', 'Капля', 'Ангел', 'Бабочка', 'Вечность', 'Надежда', 'Вера', 'Любовь', 'Мечта'],
    'Цепочки': ['Классика', 'Венецианская', 'Бисмарк', 'Панцирь', 'Змея', 'Ангел', 'Бриллиант', 'Волна', 'Река', 'Путь', 'Связь', 'Нить'],
    'Ожерелья': ['Принцесса', 'Герцогиня', 'Королева', 'Императрица', 'Чародейка', 'Фея', 'Нимфа', 'Сирена', 'Медуза', 'Наяды', 'Гера', 'Афина'],
    'Ножные браслеты': ['Афродита', 'Клеопатра', 'Нефертити', 'Венера', 'Диана', 'Юнона'],
    'Наборы': ['Романтика', 'Королевский', 'Нежность', 'Грация', 'Вечерний', 'Минимализм', 'Империя', 'Весна', 'Лето', 'Осень', 'Зима'],
}

# Изображения для категорий
CATEGORY_IMAGES = {
    'Кольца': 'https://images.unsplash.com/photo-1605100804763-247f67b3557e?w=800',
    'Серьги': 'https://images.unsplash.com/photo-1535632066927-ab7c9ab60908?w=800',
    'Браслеты': 'https://images.unsplash.com/photo-1611591437281-460bfbe1220a?w=800',
    'Кулоны': 'https://images.unsplash.com/photo-1599643478518-17488fbbcd75?w=800',
    'Цепочки': 'https://images.unsplash.com/photo-1573408301185-9146fe634ad0?w=800',
    'Ожерелья': 'https://images.unsplash.com/photo-1599643478518-17488fbbcd75?w=800',
    'Ножные браслеты': 'https://images.unsplash.com/photo-1611591437281-460bfbe1220a?w=800',
    'Наборы': 'https://images.unsplash.com/photo-1515562141207-7a88fb7ce338?w=800',
}

STONE_COLORS = {
    'Жемчуг': '#f1f5f9', 'Аметист': '#a855f7', 'Бирюза': '#06b6d4', 'Топаз': '#0ea5e9',
    'Коралл': '#f97316', 'Изумруд': '#10b981', 'Кварц': '#e2e8f0', 'Рубин': '#b91c1c',
    'Янтарь': '#f59e0b', 'Бриллиант': '#94a3b8', 'Перидот': '#84cc16', 'Гранат': '#dc2626',
    'Сапфир': '#1d4ed8', 'Яшма': '#78350f', 'Обсидиан': '#1e293b', 'Лазурит': '#1e3a8a',
    'Малахит': '#059669', 'Агат': '#64748b', 'Оникс': '#0f172a', 'Цитрин': '#fbbf24',
    'Аквамарин': '#67e8f9', 'Опал': '#fde68a',
}


def extract_stones_from_article(article):
    """Извлекает камни из артикула"""
    stones = []
    article_upper = article.upper()
    for code, stone_name in STONE_MAP.items():
        if code in article_upper:
            stones.append(stone_name)
    return stones


class Command(BaseCommand):
    help = 'Импорт товаров из silver.xlsx с правильными категориями'

    def handle(self, *args, **kwargs):
        self.stdout.write('Начало импорта...')
        
        excel_path = 'c:/Projects/Jewelry/silver.xlsx'
        if not os.path.exists(excel_path):
            self.stdout.write(self.style.ERROR('Файл не найден!'))
            return
        
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        # Очищаем базу
        from jewelry_app.models import Wishlist
        Wishlist.objects.all().delete()
        ProductImage.objects.all().delete()
        Product.objects.all().delete()
        Stone.objects.all().delete()
        Category.objects.all().delete()
        
        # Создаём категории
        categories = {}
        for cat_name in set(ARTICLE_CATEGORY_MAP.values()):
            cat, _ = Category.objects.get_or_create(name=cat_name)
            categories[cat_name] = cat
            self.stdout.write(f'Категория: {cat_name}')
        
        # Создаём камни
        stones = {}
        for code, name in STONE_MAP.items():
            stone, _ = Stone.objects.get_or_create(
                code=code,
                defaults={'name': name, 'color': STONE_COLORS.get(name, '#94a3b8')}
            )
            stones[code] = stone
            self.stdout.write(f'Камень: {name} ({code})')
        
        # Импорт товаров
        name_counters = {cat: 0 for cat in PRODUCT_NAMES.keys()}
        products_created = 0
        products_with_stones = 0
        
        for row_idx in range(2, ws.max_row + 1):
            try:
                article = str(ws.cell(row=row_idx, column=1).value or '').strip()
                if not article or article == 'None':
                    continue
                
                # Определяем категорию по артикулу
                category_name = None
                for prefix, cat_name in ARTICLE_CATEGORY_MAP.items():
                    if article.upper().startswith(prefix):
                        category_name = cat_name
                        break
                
                if not category_name:
                    category_name = 'Кольца'
                
                category = categories.get(category_name)
                if not category:
                    continue
                
                # Определяем камни из артикула
                stone_names = extract_stones_from_article(article)
                has_stones = len(stone_names) > 0
                
                if has_stones:
                    products_with_stones += 1
                
                # Генерируем название
                if category_name in PRODUCT_NAMES and name_counters[category_name] < len(PRODUCT_NAMES[category_name]):
                    product_name = PRODUCT_NAMES[category_name][name_counters[category_name]]
                    name_counters[category_name] += 1
                else:
                    product_name = f'Изделие {article}'
                
                # Читаем данные
                weight = ws.cell(row=row_idx, column=2).value
                price_thb = ws.cell(row=row_idx, column=4).value
                stock_qty = ws.cell(row=row_idx, column=7).value
                price_rub = ws.cell(row=row_idx, column=9).value
                
                # Конвертируем
                try:
                    price_rub_val = float(price_rub) if price_rub and str(price_rub) not in ['None', ''] else 0
                except:
                    price_rub_val = 0
                    
                try:
                    weight_val = float(weight) if weight and str(weight) not in ['None', ''] else None
                except:
                    weight_val = None
                    
                try:
                    price_thb_val = float(price_thb) if price_thb and str(price_thb) not in ['None', ''] else None
                except:
                    price_thb_val = None
                    
                try:
                    stock_qty_val = int(stock_qty) if stock_qty and str(stock_qty) not in ['None', ''] else 0
                except:
                    stock_qty_val = 0
                
                # Создаём продукт
                product = Product.objects.create(
                    name=product_name,
                    description=f'Серебряное украшение {product_name}' + (f' с {", ".join(stone_names)}' if has_stones else ''),
                    price=price_rub_val,
                    category=category,
                    sku=f'sku-{article}-{row_idx}',
                    article=article,
                    material='Серебро 925',
                    weight=weight_val,
                    has_stones=has_stones,
                    price_thb=price_thb_val,
                    stock_quantity=stock_qty_val,
                    is_active=True,
                    is_out_of_stock=stock_qty_val == 0,
                    is_featured=(products_created < 20),
                )
                
                # Добавляем камни через ManyToMany
                for code, stone in stones.items():
                    if code in article.upper():
                        product.stones.add(stone)
                
                # Добавляем изображение
                img_url = CATEGORY_IMAGES.get(category_name, '')
                if img_url:
                    ProductImage.objects.create(
                        product=product,
                        image_url=img_url,
                        alt_text=f'{product_name} - фото',
                        is_primary=True,
                        sort_order=0,
                    )
                
                products_created += 1
                
            except Exception as e:
                self.stdout.write(self.style.WARNING(f'Ошибка в строке {row_idx}: {e}'))
                continue
        
        self.stdout.write(self.style.SUCCESS(f'Импорт завершён! Создано {products_created} товаров'))
        self.stdout.write(self.style.SUCCESS(f'Товаров с камнями: {products_with_stones}'))
        self.stdout.write(self.style.SUCCESS(f'Всего камней: {Stone.objects.count()}'))
        
        # Вывод статистики
        self.stdout.write('\nСтатистика по категориям:')
        for cat in Category.objects.all().order_by('name'):
            count = cat.products.filter(is_active=True).count()
            self.stdout.write(f'  {cat.name}: {count} товаров')
        
        # Создаём суперпользователя с паролем "admin"
        if not User.objects.filter(username='admin').exists():
            User.objects.create_superuser('admin', 'admin@example.com', 'admin')
            self.stdout.write(self.style.SUCCESS('\nСуперпользователь создан: admin / admin'))
        else:
            admin = User.objects.get(username='admin')
            admin.set_password('admin')
            admin.save()
            self.stdout.write(self.style.SUCCESS('\nПароль администратора изменён на: admin'))
